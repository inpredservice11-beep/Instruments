"""
Модуль для экспорта данных в Excel
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime


class ExcelExporter:
    """Класс для экспорта данных в Excel"""
    
    def __init__(self):
        self.wb = None
        self.ws = None
    
    def _format_date(self, date_str):
        """Форматирование даты для отображения"""
        if not date_str:
            return ''
        
        try:
            # Пробуем разные форматы
            if ' ' in str(date_str):
                dt = datetime.strptime(str(date_str), '%Y-%m-%d %H:%M:%S')
                return dt.strftime('%d.%m.%Y %H:%M')
            else:
                dt = datetime.strptime(str(date_str), '%Y-%m-%d')
                return dt.strftime('%d.%m.%Y')
        except:
            return str(date_str)
    
    def export_history_journal(self, history_data, output_path, filter_type='Все'):
        """Экспорт журнала операций в Excel
        
        Args:
            history_data: список кортежей с данными об операциях
            output_path: путь для сохранения Excel файла
            filter_type: тип фильтра ('Все', 'Выдача', 'Возврат')
        """
        # Создаем новую книгу
        self.wb = Workbook()
        self.ws = self.wb.active
        
        # Определяем заголовок
        if filter_type == 'Все':
            title = "Журнал операций"
        elif filter_type == 'Выдача':
            title = "Журнал операций - Выдача"
        elif filter_type == 'Возврат':
            title = "Журнал операций - Возврат"
        else:
            title = "Журнал операций"
        
        # Заголовок документа
        self.ws['A1'] = title
        self.ws['A1'].font = Font(size=16, bold=True)
        self.ws['A1'].alignment = Alignment(horizontal='center')
        self.ws.merge_cells('A1:I1')
        
        # Дата формирования
        self.ws['A2'] = f"Дата формирования: {datetime.now().strftime('%d.%m.%Y %H:%M')}"
        self.ws['A2'].font = Font(size=11)
        self.ws['A2'].alignment = Alignment(horizontal='center')
        self.ws.merge_cells('A2:I2')
        
        # Заголовки таблицы
        headers = ['№', 'Тип', 'Инв. номер', 'Инструмент', 'Сотрудник', 
                  'Адрес', 'Дата операции', 'Выполнил', 'Примечание']
        
        # Стиль заголовков
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        header_font = Font(bold=True, color='FFFFFF', size=10)
        header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        # Заполняем заголовки
        for col_idx, header in enumerate(headers, 1):
            cell = self.ws.cell(row=4, column=col_idx)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
        
        # Заполняем данные
        if history_data:
            for row_idx, record in enumerate(history_data, 5):
                # Форматирование даты
                operation_date = self._format_date(record[6]) if len(record) > 6 else ''
                
                # Заполняем строку данными
                data = [
                    row_idx - 4,  # №
                    str(record[1]) if len(record) > 1 else '',  # Тип
                    str(record[2]) if len(record) > 2 else '',  # Инв. номер
                    str(record[3]) if len(record) > 3 else '',  # Инструмент
                    str(record[4]) if len(record) > 4 else '',  # Сотрудник
                    str(record[5]) if len(record) > 5 else '',  # Адрес
                    operation_date,  # Дата операции
                    str(record[7]) if len(record) > 7 else '',  # Выполнил
                    str(record[8]) if len(record) > 8 else ''   # Примечание
                ]
                
                for col_idx, value in enumerate(data, 1):
                    cell = self.ws.cell(row=row_idx, column=col_idx)
                    cell.value = value
                    cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
                    cell.font = Font(size=9)
                    
                    # Чередование цветов строк
                    if (row_idx - 4) % 2 == 0:
                        cell.fill = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')
        else:
            # Если нет данных
            self.ws['A5'] = "Нет данных для отображения"
            self.ws['A5'].alignment = Alignment(horizontal='center')
            self.ws.merge_cells('A5:I5')
        
        # Настройка ширины столбцов
        column_widths = {
            'A': 8,   # №
            'B': 12,  # Тип
            'C': 15,  # Инв. номер
            'D': 30,  # Инструмент
            'E': 25,  # Сотрудник
            'F': 35,  # Адрес
            'G': 18,  # Дата операции
            'H': 18,  # Выполнил
            'I': 30   # Примечание
        }
        
        for col_letter, width in column_widths.items():
            self.ws.column_dimensions[col_letter].width = width
        
        # Установка высоты строки заголовка
        self.ws.row_dimensions[4].height = 25
        
        # Границы для таблицы
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Применяем границы ко всем ячейкам таблицы
        if history_data:
            max_row = 4 + len(history_data)
            for row in range(4, max_row + 1):
                for col in range(1, 10):
                    self.ws.cell(row=row, column=col).border = thin_border
        
        # Итоговая информация
        if history_data:
            info_row = max_row + 2
            self.ws.cell(row=info_row, column=1).value = f"Всего записей: {len(history_data)}"
            self.ws.cell(row=info_row, column=1).font = Font(bold=True, size=10)
        
        # Сохраняем файл
        self.wb.save(output_path)






